from bs4 import BeautifulSoup
import cfscrape
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from pathlib import Path


class Scrapping:
    coins_arr = ['usd-mxn','eur-mxn','jpy-mxn','cny-mxn']
    commodities_arr = ["gold","silver","crude-oil"]
    scraper = cfscrape.create_scraper()

    def __init__(self, endpoint, endpoint_arr = [], base_url= 'https://mx.investing.com', label="span", html_class="text-2xl"):
      self.base_url = base_url
      self.endpoint = endpoint
      self.label = label
      self.html_class = html_class
      self.resultado = []
      if endpoint == 'commodities': 
        self.specific_item_arr = self.commodities_arr
      elif endpoint == 'currencies':
        self.specific_item_arr = self.coins_arr
      else:
        self.specific_item_arr = endpoint_arr
          
    def makeScrapping(self):
      for item in self.specific_item_arr:
        URL=f'{self.base_url}/{self.endpoint}/{item}'
        scraperURL = self.scraper.get(URL)
        soupURL = BeautifulSoup(scraperURL.content, "html.parser")
        valorURL = soupURL.find_all(self.label, self.html_class)
        for i in valorURL:          
          self.resultado.append(i.text)
      return self.resultado
    
    def changeStatic(self, new_value):
      self.coins_arr = self.coins_arr.append(new_value)

    def NewDoc(self):
      if os.path.isfile('Valores.xlsx'):
        cwd=Path.cwd()
      else:
        doc = Workbook()
        Mainsheet = doc.active
        Mainsheet.title="ValoresDB"
        doc.save('Valores.xlsx')
        cwd=Path.cwd()
        return cwd
    
    def InsertData(self,cwd,data,colum='A',row=1):
      doc = load_workbook(f'{cwd}/Valores.xlsx')
      Mainsheet = doc.active
      for i in data:
        Mainsheet[f'{colum}{row}']=i
        row+=1
        doc.save('Valores.xlsx')

       

